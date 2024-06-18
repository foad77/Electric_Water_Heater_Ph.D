# this file imports modules, data files and functions
import sys
import matplotlib as mpl
if len(sys.argv) >= 2:
 mpl.use('TKAgg')
from openpyxl import load_workbook
#from csv import*
from numpy import*
#import numpy as np
from pulp import*
from math import floor, ceil
#from matplotlib.pyplot import*

import tempfile
import pandas as pd, numpy as np
tempdir = tempfile.gettempdir()


import matplotlib.pyplot as plt
import matplotlib.dates as dates
import os, subprocess, platform

import time

from itertools import product
#=================================
# this is to supress all the warnings
import warnings
warnings.filterwarnings("ignore")

#===================================
# to clear the iPhython from previous runs (for spyder)
try:
    from IPython import get_ipython
    get_ipython().magic('clear')
    get_ipython().magic('reset -f')
except:
    pass
#=============================
# Importing the config file
import yaml

# Load the configuration file
with open('config.yaml', 'r') as f:
    config = yaml.safe_load(f)

# Get the variables from the configuration file
methods = config['methods']
tomorrow = config['tomorrow']
test_mode = config['test_mode']
stochastic_status = config['stochastic_status']
lookback_length = config['lookback_length']
min_slot = config['min_slot']
temp_resolution = config['temp_resolution']
Run_Time = config['Run_Time']
Optimality_gap = config['Optimality_gap']
resolution = config['resolution']
penalty_vector = config['penalty_vector']
Flate_rate_vector = config['Flate_rate_vector']
Initial_Temperature_Status = config['Initial_Temperature_Status']
t_max = config['t_max']
t_min = config['t_min']
t_min_shortfall = config['t_min_shortfall']
max_energy_input_per_hour = config['max_energy_input_per_hour']
specific_heat = config['specific_heat']
v = config['v']
heat_loss_rate = config['heat_loss_rate']
problem_formulator_status = config['problem_formulator_status']
Conservative_Penalty_Status = config['Conservative_Penalty_Status']
Ranged_Temperature_penalty_status = config['Ranged_Temperature_penalty_status']
Hourly_TS_setpoint_interval = config['Hourly_TS_setpoint_interval']
degree_of_water_per_kwh = config['degree_of_water_per_kwh']
#=============================
# it is used for high-level file and directory operation such as deleting the whole tree of files
# and directories in a folder
import shutil
start_time = time.time()

def execute_file(code): 
    
    exec(compile(open(code, "rb").read(), code, 'exec'))

## importing from 'hourly hot water' excel file
#wb = load_workbook('Input/withdrawal/{}min_usage.xlsx'.format(min_slot), read_only = True, data_only=True)
#sheet = wb['{}min_usage'.format(min_slot)]

wb = load_workbook('Input/withdrawal/usage.xlsx', read_only = True, data_only=True)
sheet = wb['{}min_usage'.format(min_slot)]
#=========================================   
# this is for price negotiating algorithm
# if len(sys.argv) == 2:
#     wb = load_workbook('Result/Communication_Link.xlsx', read_only = True, data_only=True)
#     aggregator = wb['Sheet1'] 
#========================================   
   
k_max=100;k_min=1
step=ceil((k_max)/resolution)

interval_per_hour=60/min_slot

big_temp = 60
max_energy_input_per_interval = max_energy_input_per_hour/(60/min_slot)
#==============================================================================
#### loading the results from the latest run of simulation
#==============================================================================

# indexing sets
day_set = set(); interval_set=set();interval=set()
# for past data
t_out = dict();historical_volume_withdrawn = dict();tank_temp = dict(); t_amb = dict(); t_in = dict();tomorrow_volume_withdrawn=dict();tomorrow_t_out=dict();energy_req_tomorrow=dict()
energy_req = dict(); energy_req_average = dict(); energy_req_average_normal = dict(); sim_tank_temp_normal=dict(); ts_setpoint_normal=dict(); NRG_input_normal=dict(); price_of_power = dict()
price_of_power_normal = dict(); NRG_Req_normal=dict(); E_in=dict(); t_shortfall=dict(); NRG_shortfall=dict(); NRG_withdrawn=dict(); t_tank=dict();temp_shortfall_normal=dict();NRG_shortfall_f_normal=dict()
available_solar=dict()
#   for method 7
t_target_7=dict();t_target_7_normal=dict();t_tank_7=dict();t_tank_7_normal=dict();total_cost_7=dict();total_electrical_cost_7=dict();total_shortfall_cost_7=dict();NRG_shortfall_7=dict();NRG_withdrawn_7=dict()
E_in_7=dict();t_shortfall_7=dict()
# for method 2
t_target_2=dict();total_cost_2=dict();total_electrical_cost_2=dict();total_shortfall_cost_2=dict()
# for method Du&Lu
t_tank_normal=dict();UC=dict();q_sup=dict();Cuted_Price=dict();total_cost_Du=dict();total_electrical_cost_Du=dict();total_shortfall_cost_Du=dict();t_tank_actual=dict();t_shortfall_actual=dict()
NRG_shortfall_actual=dict();NRG_withdrawn_actual=dict()
#tDictionaries for our method
total_electrical_cost_f=dict();total_shortfall_cost_f=dict();t_min_range=dict();shortfall_penalty_range=dict()
t_tank_our=dict();t_shortfall_our=dict();NRG_shortfall_our=dict();NRG_withdrawn_our=dict();E_in_our=dict();coasting=dict();at_max=dict();NRG_withdrawn_f_normal=dict()
#dictonaries for flate rate method
NRG_shortfall_flat=dict();NRG_withdrawn_flat=dict();E_in_flat=dict();t_shortfall_flat=dict();t_tank_flat=dict()
# dictionaries for average method
NRG_shortfall_WH_MILP_average=dict();t_tank_WH_MILP_average=dict();NRG_withdrawn_WH_MILP_average=dict();E_in_WH_MILP_average=dict();t_shortfall_WH_MILP_average=dict()
methods_dictionary=dict()

#=======================================================================
#    functions
#=======================================================================
# it creates a fractional range between 'start' and 'stop' with the step of 'step'.
# python by default only has integer range.
def frange(start, stop, step):
     i = start
     while i < stop:
        yield i
        i += step
# this function is used in the Du&Lu method. it takes the remaining requied energy (q_requiered) and
# the current interval (i). based on these two argumant, it finds the  energy input schedule for all
# of the remaining hours.i.e, when to turn the WH on in the following intervals.
# when the energy schedule is defined, violation of temperature deadband is checked. if it is violated for the next interval ( for example highrer than the maximum)
# the method (not this function) set the temperature to T_max and based on T_max, the requied energy is calculated which is lees than intially planed
#energy input. when requied energy of the current step is calculated (after checking with temperature violation rule), the requied energy of current
#step will be deducted from the over requied energy of the day. based on
# the new requied energy for the remaining hours, the whole process repeats until
# the schedule is obtained.
# by this function.
def U_c(q_requiered,i):
       neede_delta_t= q_requiered /max_energy_input_per_interval
       delta_t=int((q_requiered)/max_energy_input_per_interval)
       interval=(60/min_slot)*24
       for m in  range (0, int((interval+1-i))):
           Cuted_Price [m+1]=price_of_power[(int(current_day),int(i+m))]
       if delta_t >= interval-1:
           delta_t=interval-1
       Critical_Price=sorted(Cuted_Price.values())[(delta_t)]
       Price=price_of_power[(current_day,(i))]

       if (Price < Critical_Price):
           U_c = 1
       if (Price > Critical_Price):
           U_c = 0
       if (Price == Critical_Price):
          U_c =float( neede_delta_t - delta_t)
       return U_c
   
#=================================================
#======================================
# the "method" must be set to "Negotiator" if "Bidder_WH_EV.py" is being runned. to avoid this hassle,
# this piece of code obliviate that need for changing the method.
try:
    user_pattern
    methods.clear()
    try:
    
        methods=["Negotiator"]
    except:
        methods=["Negotiator_Linear"] 
        pass
        
except:
    pass
#=========================================

def status(stochastic_status,day_set,historical_volume_withdrawn,tomorrow_volume_withdrawn,tomorrow_t_out,t_out):
    global lookback_length
    if stochastic_status:

     pass
    else:
        day_set.clear()
        day_set.add(current_day)
        historical_volume_withdrawn.clear()
        t_out.clear()
        lookback_length=1
        for i in interval_set:

          historical_volume_withdrawn[current_day, i]=tomorrow_volume_withdrawn[current_day, i]
          t_out[current_day, i]=tomorrow_t_out[current_day, i]

def penalty(Ranged_Temperature_penalty_status,t_min_range,t_min):
    if Ranged_Temperature_penalty_status:
        pass
    else:
        t_min_range.clear()
        for j in range(0,resolution):
            t_min_range[j]=t_min_shortfall


def short_TS_Interval(Hourly_TS_setpoint_interval):
    global interval_per_hour
    if Hourly_TS_setpoint_interval:
       interval_per_hour=60/min_slot
    else:
       interval_per_hour=1



def init_temp(Initial_Temperature_Status,current_day):
    global temp_0_Du ,temp_0_Fixed,temp_0_Apt,temp_0_Solar
    if Initial_Temperature_Status==1:
        temp_0=t_min_range[j]
        temp_0_Du=temp_0;temp_0_Apt=temp_0;temp_0_Fixed=temp_0;temp_0_Solar=temp_0
    if Initial_Temperature_Status==2:
        temp_0=ts_setpoint[1].value()
        temp_0_Du=temp_0;temp_0_Apt=temp_0;temp_0_Fixed=temp_0;temp_0_Solar=temp_0
    if Initial_Temperature_Status==3:
        temp_0=t_min_shortfall
        temp_0_Du=temp_0;temp_0_Apt=temp_0;temp_0_Fixed=temp_0;temp_0_Solar=temp_0
    if Initial_Temperature_Status==4:
        temp_0=t_tank_our[1]
        temp_0_Apt=temp_0;temp_0_Fixed=temp_0;temp_0_Du=temp_0;temp_0_Solar=temp_0
    if Initial_Temperature_Status==5:
      if current_day==1:
          temp_0_Du=t_min_shortfall;temp_0_Apt=t_min_shortfall;temp_0_Fixed=t_min_shortfall;temp_0_Solar=t_min_shortfall      
      else:   
        try:temp_0_Du=newdf.loc[(penalty_vector[j],current_day-1),'LastTemp_Du']
        except: pass
        try:temp_0_Fixed=newdf.loc[(penalty_vector[j],current_day-1),'LastTemp_Fixed']
        except: pass
        try:temp_0_Apt=newdf.loc[(penalty_vector[j],current_day-1),'LastTemp_Apt']
        except: pass
        try:temp_0_Solar=newdf.loc[(penalty_vector[j],current_day-1),'LastTemp_Solar']
        except: pass
def formulator(problem_formulator_status):
    global Problem_formulator
    if problem_formulator_status == 1:
        Problem_formulator= 'our'
    if problem_formulator_status == 2:
        Problem_formulator= 'Pyomo'
# when want to test the model for one day in non_stochastic mode, one should a folder each time 
# to run the code. this section does it automatically. So, no need to do it each time by hand
if stochastic_status==False:
   if len(sys.argv) == 1:
    path='Result/logging/{}d_lookBack & {}m_Intrv'.format(1,min_slot)
    if os.path.exists(path)==True:
      if loop==1:  
        shutil.rmtree(path)
    else:
        pass
else:
    pass
        
        
#===========================================================
# reading the results from xlsx file. if it exist, only read it. if not, then creat one
days=[]
for i in range(1,365):
        days.append(i)
iterables = [penalty_vector,days]
cost_type=["C", "Dis"]
indexing=pd.MultiIndex.from_product(iterables, names=['penalty','day'])
names_combination=[]
for word, letter, hours in product(["Exp", "Act"],cost_type, methods):
    string='{}_{}_{}'.format(word, letter, hours)
    names_combination.append(string)
for  hours in methods:
    srts='LastTemp_{}'.format(hours)
    names_combination.append(srts)
check_if_runned=555    


if stochastic_status:condition='stochastic'
else:condition= 'perfect_forecast'
if len(sys.argv) == 4: lookback_length=int(sys.argv[3])
#os.path.isfile(os.path.join(flate_rate_path, name)

#os.path.exists(DIR)==True: 
#os.path.isfile(os.path.join(DIR, name))])
#os.path.isfile(current_file)==True:
Current_Path='Result/logging/{}/{}d_lookBack_{}m_Intrv'.format(condition,lookback_length,min_slot)
Current_File= os.path.join(Current_Path, 'file.xlsx')
try: 
    newdf = pd.read_excel(Current_File) 
    #newdf.reset_index(inplace=True)
    newdf.set_index(['penalty','day'],inplace=True)
    for t in names_combination:
        if t in newdf.columns:
            pass
        else:
            newdf[t]=check_if_runned
    
except:
    newdf= pd.DataFrame(index=indexing,columns=names_combination)
    newdf.fillna(check_if_runned,inplace=True)
    try:
       os.makedirs(Current_Path)
    except:
        pass
t=2


#=======================
tempDf= pd.DataFrame()
#========================

# this function is used to generate actual result. the input would be termostat setpoint that is generated by each method (MILP,Apt,Goh,...).
#then, this function generates actual temperature based on energy requied of tomorrow.
OneDayDF= pd.DataFrame()
def Actual_Result_Generator(temperature_setpoints,temp_0):
   global OneDayDF
   Ew1=dict();Ew2=dict();t_tank_conservative=dict();E_error=dict();NRG_short=dict();t_Actual=dict();E_Input=dict();t_setpoint=dict()
   dictionaries= [Ew1,Ew2,E_error,NRG_short,t_Actual,NRG_withdrawn,E_Input,t_shortfall]
   for d in dictionaries:
      d.clear()
   #for b in temperature_setpoints.keys():
   for b in sorted(interval_set):
        NRG_short[b]=0
        if b == 1:
            i_prev = max(interval_set)
        else:
            i_prev = b - 1            
        t_Actual[24*(60/min_slot)]=temp_0
        t_setpoint[b]=temperature_setpoints[b]
        t_Actual[b]=t_setpoint[b]
        NRG_withdrawn[b]=energy_req_tomorrow[b]
#        t= t_Actual[b]*(1+heat_loss_rate) - heat_loss_rate * t_amb[current_day,b] - t_Actual[i_prev]
#        k=  degree_of_water_per_kwh * (t)
#        E_Input[b]=k+NRG_withdrawn[b]+0.00001
        E_Input[b]=degree_of_water_per_kwh * (t_Actual[b]*(1+heat_loss_rate) - heat_loss_rate * t_amb[current_day,b] - t_Actual[i_prev])+NRG_withdrawn[b]+0.00001

        Ew1[b]=1; Ew2[b]=0
        
        if (E_Input[b] >= max_energy_input_per_interval):
          E_Input[b]=max_energy_input_per_interval
          t_Actual[b]=(1/(1+heat_loss_rate))*(((E_Input[b]-NRG_withdrawn[b])/degree_of_water_per_kwh)+t_amb[current_day,b]*heat_loss_rate+t_Actual[i_prev])
          
          while (abs(Ew1[b]-Ew2[b]) >= 0.001):      
                  
            t_shortfall[b] = t_min_shortfall - t_Actual[b]+0.0000001
            if (t_shortfall[b] <0):
                t_shortfall[b] =0
            Ew1[b] = energy_req_tomorrow[b] * t_shortfall[b] / (t_min_shortfall - t_in[current_day])+0.0000001
            k=  degree_of_water_per_kwh * (t_Actual[b]*(1+heat_loss_rate) - heat_loss_rate * t_amb[current_day,b] - t_Actual[i_prev])+0.0000001
            Ew2[b]=k+energy_req_tomorrow[b]-max_energy_input_per_interval

            NRG_short[b]=Ew1[b]
            t_Actual[b]=t_Actual[b]+0.001
            E_error[b]=abs(Ew1[b]-Ew2[b])

        if (E_Input[b] <=0):

          while (abs(Ew1[b]-Ew2[b]) >= 0.001):
            
            Ew1[b]=degree_of_water_per_kwh*(t_Actual[b]*(-1-heat_loss_rate)+ t_Actual[i_prev] + heat_loss_rate * t_amb[current_day,b])+0.0000001        
            t_shortfall[b] = t_min_shortfall - t_Actual[b]+0.0000001
            if (t_shortfall[b] <0):
                t_shortfall[b] =0
            NRG_short[b] = energy_req_tomorrow[b] * t_shortfall[b] / (t_min_shortfall - t_in[current_day])
            Ew2[b] = energy_req_tomorrow[b] - NRG_short[b]

            k=  degree_of_water_per_kwh * (t_Actual[b]*(1+heat_loss_rate) - heat_loss_rate * t_amb[current_day,b] - t_Actual[i_prev])
            E_Input[b]=k+Ew1[b]+0.0000001
            if (E_Input[b] <= -0.1):
                 Ew1[b]=1
                 Ew2[b]=0
    
            t_Actual[b]=t_Actual[b] +0.001
            E_error[b]=abs(Ew1[b]-Ew2[b])
            
        if (E_Input[b] >=0.1) and (E_Input[b] <=max_energy_input_per_interval-0.1) and (t_Actual[b] <= t_min_shortfall-0.2) :
            t_shortfall[b] = t_min_shortfall - t_Actual[b]+0.0000001
            NRG_short[b] = energy_req_tomorrow[b] * t_shortfall[b] / (t_min_shortfall - t_in[current_day])
            k=  degree_of_water_per_kwh * (t_Actual[b]*(1+heat_loss_rate) - heat_loss_rate * t_amb[current_day,b] - t_Actual[i_prev])+0.0000001
            E_Input[b]=k+energy_req_tomorrow[b]-NRG_short[b]+0.000001 
        OneDayDF.loc[b,'Act_temp_{}'.format(methods_series[m])]= t_Actual[b] 
        OneDayDF.loc[b,'set_temp_{}'.format(methods_series[m])]= temperature_setpoints[b]
        OneDayDF.loc[b,'NRG_in_{}'.format(methods_series[m])]= E_Input[b]

            
   electrical_cost=sum(price_of_power[current_day,b] * E_Input[b]  for b in interval_set)                  
   shortfall_cost = sum(NRG_short[b] for b in interval_set)
   #return   electrical_cost,  shortfall_cost
   
   Actual_dict={}
   Actual_dict={'Act_Dis_{}'.format(methods_series[m]):shortfall_cost,'Act_C_{}'.format(methods_series[m]):electrical_cost,'LastTemp_{}'.format(methods_series[m]):t_Actual[max(interval_set)]}
   return      Actual_dict      
    
    
    
    
    

    


         
    
    
    
    